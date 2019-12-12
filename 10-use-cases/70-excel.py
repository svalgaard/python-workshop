#! /usr/bin/env python3
#
# Requires openpyxl
#
import sys
import openpyxl


def main():
    global cell, row
    filename = sys.argv[1]

    print(f'Opening xlsx file {filename}')
    # Without the data_only flag, we get formulas and not values
    wb = openpyxl.load_workbook(filename, data_only=True)

    print(f"All sheet names {wb.sheetnames}")

    for sheetname in wb.sheetnames:
        print(f"Current sheet name is {sheetname}")
        ws = wb[sheetname]

        for row in ws.rows:
            for cell in row:
                print(f'Cell {cell.coordinate} has value {cell.value}')
            print()


if __name__ == '__main__':
    main()
